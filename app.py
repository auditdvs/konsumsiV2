import io
import re
from difflib import SequenceMatcher

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Breakdown Biaya Konsumsi", layout="wide")

FONT_NAME = "Arial"

# =========================================================
# 1. PARSER FILE LAPORAN
# =========================================================
#
# Mendukung dua format:
# A. Format lama "TRANSACTION LISTING BY ACCOUNTS" (blok per Account No,
#    diakhiri baris TOTAL).
# B. Format baru: tabel flat dengan header
#    VOUCHER NO. | TRANS. DATE | ENTRY DATE | DESCRIPTION | DEBIT | CREDIT
#    tanpa blok Account No / TOTAL. Parser mendeteksi baris header ini
#    secara otomatis lalu membaca seluruh baris di bawahnya sebagai data.

def load_raw_table(uploaded_file):
    """Coba beberapa cara baca file: html (xls export), excel asli, atau csv."""
    raw_bytes = uploaded_file.read()

    try:
        tables = pd.read_html(io.BytesIO(raw_bytes))
        tables = sorted(tables, key=lambda t: t.shape[0], reverse=True)
        if tables and tables[0].shape[0] > 5:
            return tables[0]
    except Exception:
        pass

    try:
        return pd.read_excel(io.BytesIO(raw_bytes), header=None)
    except Exception:
        pass

    try:
        return pd.read_csv(io.BytesIO(raw_bytes), header=None)
    except Exception:
        pass

    return None


def _parse_date(val) -> "pd.Timestamp":
    """Parse tanggal secara fleksibel: handle string DD/MM/YYYY, auto-parsed datetime, dll."""
    try:
        s = str(val).strip()
        if s in ("nan", "", "NaT", "None", "nat"):
            return pd.NaT
        # Coba parse langsung (untuk datetime object atau ISO format)
        return pd.to_datetime(s, dayfirst=True)
    except Exception:
        return pd.NaT


def _parse_num(val) -> float:
    """Parse angka dari cell; handle koma ribuan, string kosong, dll."""
    try:
        s = str(val).replace(",", "").replace(" ", "").strip()
        if s in ("nan", "", "-", "NaT", "None"):
            return 0.0
        return float(s)
    except Exception:
        return 0.0


def _find_header_row(raw_df: pd.DataFrame, max_scan: int = 15):
    """
    Cari baris header tabel flat (mengandung 'VOUCHER' dan 'DEBIT') di antara
    max_scan baris pertama. Return (index_baris, list_nama_kolom_upper) atau
    (None, None) kalau tidak ditemukan.
    """
    n_scan = min(max_scan, len(raw_df))
    for i in range(n_scan):
        row_vals = [str(v).strip().upper() for v in raw_df.iloc[i].tolist()]
        if any("VOUCHER" in v for v in row_vals) and any("DEBIT" in v for v in row_vals):
            return i, row_vals
    return None, None


def _parse_flat_table(raw_df: pd.DataFrame, header_idx: int, header_vals: list) -> pd.DataFrame:
    """
    Parser untuk format baru: tabel flat tanpa blok Account No / TOTAL.
    Kolom diidentifikasi berdasarkan nama header, bukan posisi tetap, supaya
    tetap tahan kalau urutan kolom sedikit berubah.
    """
    col_map = {name: j for j, name in enumerate(header_vals)}

    def find_col(*keys):
        for name, j in col_map.items():
            if any(key in name for key in keys):
                return j
        return None

    idx_voucher = find_col("VOUCHER")
    idx_trans = find_col("TRANS")
    idx_entry = find_col("ENTRY")
    idx_desc = find_col("DESCRIPTION", "DESKRIPSI", "KETERANGAN")
    idx_debit = find_col("DEBIT")
    idx_credit = find_col("CREDIT")

    records = []
    n_cols = raw_df.shape[1]

    for i in range(header_idx + 1, len(raw_df)):
        c0 = raw_df.iat[i, idx_voucher] if idx_voucher is not None and idx_voucher < n_cols else None
        c0_str = str(c0).strip() if c0 is not None else ""

        if c0_str in ("", "nan", "None"):
            continue
        if re.match(r"^\s*(TOTAL|ENDING BALANCE|GRAND TOTAL)", c0_str, re.IGNORECASE):
            continue

        tanggal = _parse_date(raw_df.iat[i, idx_trans]) if idx_trans is not None else pd.NaT
        if pd.isna(tanggal):
            # baris tanpa tanggal valid (mis. baris kosong / catatan kaki) dilewati
            continue

        entry_date = _parse_date(raw_df.iat[i, idx_entry]) if idx_entry is not None else pd.NaT
        deskripsi = str(raw_df.iat[i, idx_desc]) if idx_desc is not None else ""
        debit = _parse_num(raw_df.iat[i, idx_debit]) if idx_debit is not None else 0.0
        credit = _parse_num(raw_df.iat[i, idx_credit]) if idx_credit is not None else 0.0

        records.append(
            dict(
                akun=None,
                voucher=c0_str,
                tanggal=tanggal,
                entry_date=entry_date,
                deskripsi=deskripsi,
                debit=debit,
                credit=credit,
            )
        )

    return pd.DataFrame(records)


def _parse_block_table(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    Parser format lama "TRANSACTION LISTING BY ACCOUNTS": baris data berada
    di antara baris 'Account No. ...' dan baris 'TOTAL ...'.
    Kolom (0-indexed): 0=voucher, 1=trans date, 2=entry date, 3=deskripsi,
                       4=debit, 5=credit.
    """
    records = []
    in_data = False
    current_account = None
    n_cols = raw_df.shape[1]

    for i in range(len(raw_df)):
        c0 = str(raw_df.iat[i, 0]) if n_cols > 0 else ""

        if "Account No" in c0:
            current_account = c0.split(",")[1].strip() if "," in c0 else c0
            in_data = True
            continue

        if re.match(r"^\s*(TOTAL|ENDING BALANCE|GRAND TOTAL)", c0, re.IGNORECASE):
            in_data = False
            continue

        if not in_data:
            continue

        if c0.strip() in ("", "nan"):
            continue

        tanggal = _parse_date(raw_df.iat[i, 1])
        if pd.isna(tanggal):
            continue

        entry_date = _parse_date(raw_df.iat[i, 2]) if n_cols > 2 else pd.NaT
        deskripsi = str(raw_df.iat[i, 3]) if n_cols > 3 else ""
        debit = _parse_num(raw_df.iat[i, 4]) if n_cols > 4 else 0.0
        credit = _parse_num(raw_df.iat[i, 5]) if n_cols > 5 else 0.0

        records.append(
            dict(
                akun=current_account,
                voucher=c0,
                tanggal=tanggal,
                entry_date=entry_date,
                deskripsi=deskripsi,
                debit=debit,
                credit=credit,
            )
        )

    return pd.DataFrame(records)


def parse_report(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    Deteksi format sumber secara otomatis:
    - Kalau ketemu baris header 'VOUCHER ... DEBIT ...' -> format tabel flat baru.
    - Kalau tidak -> fallback ke format lama berbasis blok Account No / TOTAL.
    """
    header_idx, header_vals = _find_header_row(raw_df)
    if header_idx is not None:
        result = _parse_flat_table(raw_df, header_idx, header_vals)
        if not result.empty:
            return result

    return _parse_block_table(raw_df)



# =========================================================
# 2. KATEGORISASI TRANSAKSI (berbasis klausa "dan" / koma)
# =========================================================

def split_clauses(desc: str):
    return re.split(r"\bdan\b|,", desc.lower())


def _is_mini_training(text: str) -> bool:
    """Deteksi 'mini training' dengan toleransi typo (mis. 'mini treaning')."""
    if "mini train" in text:
        return True
    # Cek setiap pasangan kata berurutan dengan fuzzy ratio >= 0.80
    words = text.split()
    for i in range(len(words) - 1):
        candidate = f"{words[i]} {words[i + 1]}"
        if SequenceMatcher(None, candidate, "mini training").ratio() >= 0.80:
            return True
    return False


def categorize(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, r in df.iterrows():
        desc = str(r["deskripsi"])
        clauses = split_clauses(desc)

        beras_kg = 0
        beras_mentioned = False
        gula_kg = 0
        gula_mentioned = False
        galon_baru = 0
        isi_ulang = 0
        kopi = False
        teh = False
        ktg = False
        jumsih = False
        mini_training = False
        bukber = False

        kopi_qty = 0
        kopi_no_qty_flag = False
        teh_qty = 0
        teh_no_qty_flag = False

        for c in clauses:
            nums = re.findall(r"\d+", c)
            qty = int(nums[0]) if nums else 1

            # Deteksi galon: tangkap typo "galo" (missing 'n') dengan \bgalon?\b
            _has_galon = bool(re.search(r"\bgalon?\b", c)) or "gallon" in c

            is_refill = "isi ulang" in c and (
                _has_galon or "aqua" in c or "asli" in c
            )
            is_new_galon = (not is_refill) and (
                _has_galon
                or ("aqua" in c and "asli" in c)
                or ("air" in c and _has_galon)
            )

            if is_refill:
                isi_ulang += qty
            elif is_new_galon:
                galon_baru += qty

            if "beras" in c:
                beras_mentioned = True
                m = re.search(r"(\d+)\s*kg", c)
                if m:
                    beras_kg += int(m.group(1))

            if "gula" in c:
                gula_mentioned = True
                m2 = re.search(r"(\d+)\s*kg", c)
                if m2:
                    gula_kg += int(m2.group(1))

            if "kopi" in c:
                kopi = True
                # Cari qty: "N [unit] kopi" atau "kopi N [unit]"
                _UNIT_KOPI = r"(?:bungkus|bks|sachet|pcs|pak|pack|buah|unit|kaleng|kotak)?"
                mk = re.search(rf"(\d+)\s*{_UNIT_KOPI}\s*kopi", c) or re.search(rf"kopi\s*(\d+)\s*{_UNIT_KOPI}", c)
                if mk:
                    kopi_qty += int(mk.group(1))
                else:
                    kopi_no_qty_flag = True

            if "teh" in c:
                teh = True
                # Cari qty: "N [unit] teh" atau "teh N [unit]"
                _UNIT_TEH = r"(?:kantung|kantong|pcs|sachet|pak|pack|bungkus|bks|buah|unit|kaleng|kotak|box)?"
                mt = re.search(rf"(\d+)\s*{_UNIT_TEH}\s*teh", c) or re.search(rf"teh\s*(\d+)\s*{_UNIT_TEH}", c)
                if mt:
                    teh_qty += int(mt.group(1))
                else:
                    teh_no_qty_flag = True

            if "kopi" in c or "gula" in c or "teh" in c:
                ktg = True
            if "jumsih" in c or "jumsi" in c:
                jumsih = True
            if "mini train" in c or _is_mini_training(c):
                mini_training = True
            if "bukber" in c or "buka bersama" in c:
                bukber = True

        rows.append(
            dict(
                voucher=r["voucher"],
                tanggal=r["tanggal"],
                entry_date=r["entry_date"],
                deskripsi=desc,
                debit=r["debit"],
                beras_kg=beras_kg,
                beras_no_qty=(beras_mentioned and beras_kg == 0),
                gula_kg=gula_kg,
                gula_no_qty=(gula_mentioned and gula_kg == 0),
                galon_baru=galon_baru,
                isi_ulang=isi_ulang,
                kopi=kopi,
                kopi_qty=kopi_qty,
                kopi_no_qty=kopi_no_qty_flag,
                teh=teh,
                teh_qty=teh_qty,
                teh_no_qty=teh_no_qty_flag,
                ktg=ktg,
                jumsih=jumsih,
                mini_training=mini_training,
                bukber=bukber,
            )
        )

    out = pd.DataFrame(rows)
    out["bulan"] = out["tanggal"].dt.to_period("M")
    return out


def estimate_unit_prices(cat: pd.DataFrame):
    """Estimasi harga satuan dari transaksi yang isinya hanya 1 jenis item."""

    def pure(mask_field):
        others = ["beras_kg", "galon_baru", "isi_ulang"]
        others = [o for o in others if o != mask_field]
        m = (cat[mask_field] > 0)
        for o in others:
            m &= cat[o] == 0
        m &= ~(cat.ktg | cat.jumsih | cat.mini_training | cat.bukber)
        return cat[m]

    pb = pure("beras_kg")
    pg = pure("galon_baru")
    pi = pure("isi_ulang")

    harga_beras = (pb.debit.sum() / pb.beras_kg.sum()) if pb.beras_kg.sum() else 0
    harga_galon = (pg.debit.sum() / pg.galon_baru.sum()) if pg.galon_baru.sum() else 0
    harga_isi_ulang = (pi.debit.sum() / pi.isi_ulang.sum()) if pi.isi_ulang.sum() else 0

    return harga_beras, harga_galon, harga_isi_ulang


def allocate_costs(cat: pd.DataFrame, harga_beras, harga_galon, harga_isi_ulang):
    """
    Untuk transaksi gabungan (mis. 'beras 20 kg dan aqua asli 4 buah'), kurangi
    dulu estimasi biaya item yang harganya diketahui, sisanya (remainder)
    dianggap sebagai biaya kopi/gula/teh atau kegiatan (jumsih/mini training/bukber)
    yang disebut di transaksi yang sama.
    """
    df = cat.copy()
    df["biaya_beras"] = df.beras_kg * harga_beras
    df["biaya_galon_baru"] = df.galon_baru * harga_galon
    df["biaya_isi_ulang"] = df.isi_ulang * harga_isi_ulang
    df["biaya_diketahui"] = df.biaya_beras + df.biaya_galon_baru + df.biaya_isi_ulang
    df["sisa"] = (df.debit - df.biaya_diketahui).clip(lower=0)

    df["biaya_ktg"] = 0.0
    df["biaya_jumsih"] = 0.0
    df["biaya_mini_training"] = 0.0
    df["biaya_bukber"] = 0.0
    df["biaya_lainnya"] = 0.0

    for i, r in df.iterrows():
        sisa = r["sisa"]
        flags = [f for f in ["ktg", "jumsih", "mini_training", "bukber"] if r[f]]
        if not flags:
            df.at[i, "biaya_lainnya"] = sisa
            continue
        share = sisa / len(flags)
        for f in flags:
            df.at[i, f"biaya_{f}"] = share

    return df


# =========================================================
# 3. EXPORT EXCEL MULTI SHEET
# =========================================================

HEADER_FONT = Font(name=FONT_NAME, bold=True)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12)
BODY_FONT = Font(name=FONT_NAME)
RP_FORMAT = "#,##0"


def _write_table(ws, start_row, headers, rows, currency_cols=None):
    """Tulis satu tabel mulai dari start_row. Return baris berikutnya yang kosong."""
    currency_cols = currency_cols or []
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = HEADER_FONT
    r = start_row + 1
    for row_vals in rows:
        for j, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = BODY_FONT
            if j in currency_cols:
                c.number_format = RP_FORMAT
        r += 1
    for j in range(1, len(headers) + 1):
        col_letter = get_column_letter(j)
        max_len = len(str(headers[j - 1]))
        for row_vals in rows:
            if j - 1 < len(row_vals):
                max_len = max(max_len, len(str(row_vals[j - 1])))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)
    return r + 1


def _write_title(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    return row + 1


def build_excel(alloc: pd.DataFrame, harga_beras, harga_galon, harga_isi_ulang) -> bytes:
    wb = Workbook()

    alloc = alloc.assign(bulan_str=alloc["bulan"].astype(str))

    # --- Sheet 1: Ringkasan ---
    ws = wb.active
    ws.title = "Ringkasan"

    # Mask transaksi lainnya (tidak masuk kategori manapun) - sama persis dgn Sheet 8
    mask_lainnya_r = (
        (alloc["galon_baru"] == 0) & (alloc["isi_ulang"] == 0) &
        (alloc["beras_kg"] == 0) & (~alloc["beras_no_qty"]) &
        (~alloc["ktg"]) & (~alloc["jumsih"]) &
        (~alloc["mini_training"]) & (~alloc["bukber"])
    )
    alloc["debit_lainnya"] = alloc["debit"].where(mask_lainnya_r, 0)

    ring = alloc.groupby("bulan_str").agg(
        Kopi_Gula_Teh=("biaya_ktg", "sum"),
        Beras=("biaya_beras", "sum"),
        Galon_Baru=("biaya_galon_baru", "sum"),
        Isi_Ulang=("biaya_isi_ulang", "sum"),
        Jumsih=("biaya_jumsih", "sum"),
        Mini_Training=("biaya_mini_training", "sum"),
        Bukber=("biaya_bukber", "sum"),
        Lainnya=("debit_lainnya", "sum"),
        Total_Debit_Asli=("debit", "sum"),
    ).reset_index()
    ring["Total_Estimasi"] = ring[[
        "Kopi_Gula_Teh", "Beras", "Galon_Baru", "Isi_Ulang", "Jumsih",
        "Mini_Training", "Bukber", "Lainnya",
    ]].sum(axis=1)
    ring["Selisih"] = ring["Total_Estimasi"] - ring["Total_Debit_Asli"]

    headers = ["Bulan", "Kopi/Gula/Teh (Rp)", "Beras (Rp)", "Galon Baru (Rp)",
               "Isi Ulang (Rp)", "Jumsih (Rp)", "Mini Training (Rp)", "Bukber (Rp)",
               "Lainnya (Rp)", "Total Debit Asli (Rp)", "Total Estimasi (Rp)", "Selisih (Rp)"]
    rows = ring[["bulan_str", "Kopi_Gula_Teh", "Beras", "Galon_Baru", "Isi_Ulang",
                 "Jumsih", "Mini_Training", "Bukber", "Lainnya", "Total_Debit_Asli",
                 "Total_Estimasi", "Selisih"]].round(0).values.tolist()
    r = _write_title(ws, 1, "Ringkasan biaya konsumsi per bulan")
    r += 1
    r = _write_table(ws, r, headers, rows, currency_cols=list(range(2, 13)))
    r += 1
    ws.cell(row=r, column=1, value="Catatan:").font = HEADER_FONT
    r += 1
    ws.cell(row=r, column=1,
            value="Angka per kategori adalah estimasi. Kolom 'Lainnya' berisi nominal debit asli "
                  "transaksi yang tidak terdeteksi masuk kategori manapun.").font = BODY_FONT
    r += 1
    ws.cell(row=r, column=1,
            value=f"Estimasi harga per kg beras: Rp {harga_beras:,.0f}".replace(",", ".")).font = BODY_FONT
    r += 1
    ws.cell(row=r, column=1,
            value=f"Estimasi harga per galon aqua asli (baru): Rp {harga_galon:,.0f}".replace(",", ".")).font = BODY_FONT
    r += 1
    ws.cell(row=r, column=1,
            value=f"Estimasi harga per isi ulang galon: Rp {harga_isi_ulang:,.0f}".replace(",", ".")).font = BODY_FONT

    # --- Sheet 2: Galon ---
    ws2 = wb.create_sheet("Galon")
    galon_df = alloc[(alloc["galon_baru"] > 0) | (alloc["isi_ulang"] > 0)].copy()

    BULAN_ID_G = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember",
    }

    galon_df["bulan_label"] = galon_df["tanggal"].apply(lambda d: f"{BULAN_ID_G[d.month]} {d.year}")
    galon_df["trans_date_str"] = galon_df["tanggal"].dt.strftime("%d/%m/%Y")
    galon_df["entry_date_str"] = galon_df["entry_date"].apply(
        lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) else ""
    ) if "entry_date" in galon_df.columns else ""

    galon_rows = []
    for _, row in galon_df.iterrows():
        galon_rows.append([
            row["voucher"],
            row["deskripsi"],
            row["trans_date_str"],
            row["entry_date_str"],
            row["bulan_label"],
            row["galon_baru"],
            row["isi_ulang"],
            row["debit"]
        ])

    r = _write_title(ws2, 1, "Ringkasan Galon per Bulan")
    r += 1
    if galon_df.empty:
        ws2.cell(row=r, column=1, value="Tidak ada transaksi galon.").font = BODY_FONT
        r += 1
    else:
        galon_df["_sort_key"] = galon_df["tanggal"].dt.year * 100 + galon_df["tanggal"].dt.month
        g_sum = galon_df.groupby(["bulan_label", "_sort_key"]).agg(
            g_asli=("galon_baru", "sum"),
            g_isiulang=("isi_ulang", "sum"),
            nominal=("debit", "sum")
        ).reset_index().sort_values("_sort_key")

        summary_rows = []
        for _, row in g_sum.iterrows():
            summary_rows.append([
                row["bulan_label"],
                row["g_asli"],
                row["g_isiulang"],
                row["nominal"]
            ])
        r = _write_table(
            ws2, r,
            ["Periode", "Galon (asli)", "Galon (isi ulang)", "Nominal (Rp)"],
            summary_rows,
            currency_cols=[4]
        )

    r += 1
    r = _write_title(ws2, r, "Rincian Transaksi Galon (Baru & Isi Ulang)")
    r += 1
    if galon_df.empty:
        ws2.cell(row=r, column=1, value="Tidak ada rincian transaksi galon.").font = BODY_FONT
        r += 1
    else:
        r = _write_table(
            ws2, r,
            ["No Voucher", "Keterangan", "TRANS DATE.", "ENTRY DATE", "Bulan - Tahun", "Galon (asli)", "Galon (isiulang)", "Nominal (Rp)"],
            galon_rows,
            currency_cols=[8]
        )

    # --- Sheet 3: Beras ---
    ws3 = wb.create_sheet("Beras")
    beras_df = alloc[(alloc["beras_kg"] > 0) | alloc["beras_no_qty"]].copy()

    BULAN_ID_B = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember",
    }

    beras_df["bulan_label"] = beras_df["tanggal"].apply(lambda d: f"{BULAN_ID_B[d.month]} {d.year}")
    beras_df["trans_date_str"] = beras_df["tanggal"].dt.strftime("%d/%m/%Y")
    beras_df["entry_date_str"] = beras_df["entry_date"].apply(
        lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) else ""
    ) if "entry_date" in beras_df.columns else ""

    beras_rows = []
    for _, row in beras_df.iterrows():
        beras_rows.append([
            row["voucher"],
            row["deskripsi"],
            row["trans_date_str"],
            row["entry_date_str"],
            row["bulan_label"],
            row["beras_kg"] if not row["beras_no_qty"] else 0,
            row["debit"]
        ])

    r = _write_title(ws3, 1, "Ringkasan Beras per Bulan")
    r += 1
    if beras_df.empty:
        ws3.cell(row=r, column=1, value="Tidak ada transaksi beras.").font = BODY_FONT
        r += 1
    else:
        beras_df["_sort_key"] = beras_df["tanggal"].dt.year * 100 + beras_df["tanggal"].dt.month
        b_sum = beras_df.groupby(["bulan_label", "_sort_key"]).agg(
            total_kg=("beras_kg", "sum"),
            nominal=("debit", "sum")
        ).reset_index().sort_values("_sort_key")

        summary_rows_b = []
        for _, row in b_sum.iterrows():
            summary_rows_b.append([
                row["bulan_label"],
                row["total_kg"],
                row["nominal"]
            ])
        r = _write_table(
            ws3, r,
            ["Periode", "Beras (kg)", "Nominal (Rp)"],
            summary_rows_b,
            currency_cols=[3]
        )

    r += 1
    r = _write_title(ws3, r, "Rincian Transaksi Beras")
    r += 1
    if beras_df.empty:
        ws3.cell(row=r, column=1, value="Tidak ada rincian transaksi beras.").font = BODY_FONT
        r += 1
    else:
        r = _write_table(
            ws3, r,
            ["No Voucher", "Keterangan", "TRANS DATE.", "ENTRY DATE", "Bulan - Tahun", "Beras (kg)", "Nominal (Rp)"],
            beras_rows,
            currency_cols=[7]
        )

    # --- Sheet 4: Gula ---
    ws4 = wb.create_sheet("Gula")
    gu1 = alloc.groupby("bulan_str").agg(Total_Kg=("gula_kg", "sum")).reset_index()
    gu2 = alloc[alloc.gula_no_qty][["voucher", "tanggal", "deskripsi", "debit"]].copy()
    gu2["tanggal"] = gu2["tanggal"].dt.strftime("%d/%m/%Y")

    r = _write_title(ws4, 1, "Total kg gula per bulan (dari transaksi yang mencantumkan kg)")
    r += 1
    r = _write_table(
        ws4, r, ["Bulan", "Total Kg"],
        gu1[["bulan_str", "Total_Kg"]].values.tolist(),
    )
    r += 1
    r = _write_title(ws4, r, "Transaksi gula yang tidak tercatat jumlah kg-nya")
    r += 1
    if len(gu2) == 0:
        ws4.cell(row=r, column=1, value="Tidak ada transaksi gula tanpa keterangan kg.").font = BODY_FONT
    else:
        r = _write_table(
            ws4, r, ["Voucher No.", "Tanggal", "Deskripsi", "Nominal (Rp)"],
            gu2[["voucher", "tanggal", "deskripsi", "debit"]].values.tolist(),
            currency_cols=[4],
        )

    # Mapping bulan angka -> nama Indonesia
    BULAN_ID = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember",
    }

    def fmt_bulan(period_str: str) -> str:
        """'2026-01' -> 'Januari - 2026'"""
        try:
            parts = period_str.split("-")
            return f"{BULAN_ID[int(parts[1])]} - {parts[0]}"
        except Exception:
            return period_str

    # --- Sheet 5: Kopi ---
    ws5 = wb.create_sheet("Kopi")
    kopi_all = alloc[alloc["kopi"]].copy()
    kopi_all["tanggal_str"] = kopi_all["tanggal"].dt.strftime("%d/%m/%Y")

    # Tabel qty per bulan
    kopi_qty_per_bulan = (
        kopi_all[kopi_all["kopi_qty"] > 0]
        .groupby("bulan_str")["kopi_qty"].sum()
        .reset_index()
    )
    kopi_qty_per_bulan["Bulan - Tahun"] = kopi_qty_per_bulan["bulan_str"].apply(fmt_bulan)

    # Voucher tanpa qty
    kopi_no_qty_df = kopi_all[kopi_all["kopi_no_qty"]][["voucher", "tanggal_str", "deskripsi", "debit"]]

    r = _write_title(ws5, 1, "Kopi - Qty (bungkus/sachet) per bulan")
    r += 1
    if kopi_qty_per_bulan.empty:
        ws5.cell(row=r, column=1, value="Tidak ada transaksi kopi dengan informasi qty.").font = BODY_FONT
        r += 1
    else:
        r = _write_table(
            ws5, r, ["Bulan - Tahun", "Total Qty (bungkus/sachet)"],
            kopi_qty_per_bulan[["Bulan - Tahun", "kopi_qty"]].values.tolist(),
        )

    if not kopi_no_qty_df.empty:
        r += 1
        r = _write_title(ws5, r, "Transaksi kopi tanpa informasi jumlah (voucher acuan)")
        r += 1
        r = _write_table(
            ws5, r, ["Voucher No.", "Tanggal", "Deskripsi", "Nominal (Rp)"],
            kopi_no_qty_df[["voucher", "tanggal_str", "deskripsi", "debit"]].values.tolist(),
            currency_cols=[4],
        )

    # --- Sheet 6: Teh ---
    ws6 = wb.create_sheet("Teh")
    teh_all = alloc[alloc["teh"]].copy()
    teh_all["tanggal_str"] = teh_all["tanggal"].dt.strftime("%d/%m/%Y")

    # Tabel qty per bulan
    teh_qty_per_bulan = (
        teh_all[teh_all["teh_qty"] > 0]
        .groupby("bulan_str")["teh_qty"].sum()
        .reset_index()
    )
    teh_qty_per_bulan["Bulan - Tahun"] = teh_qty_per_bulan["bulan_str"].apply(fmt_bulan)

    # Voucher tanpa qty
    teh_no_qty_df = teh_all[teh_all["teh_no_qty"]][["voucher", "tanggal_str", "deskripsi", "debit"]]

    r = _write_title(ws6, 1, "Teh - Qty (kantung/kantong/pcs) per bulan")
    r += 1
    if teh_qty_per_bulan.empty:
        ws6.cell(row=r, column=1, value="Tidak ada transaksi teh dengan informasi qty.").font = BODY_FONT
        r += 1
    else:
        r = _write_table(
            ws6, r, ["Bulan - Tahun", "Total Qty (kantung/pcs)"],
            teh_qty_per_bulan[["Bulan - Tahun", "teh_qty"]].values.tolist(),
        )

    if not teh_no_qty_df.empty:
        r += 1
        r = _write_title(ws6, r, "Transaksi teh tanpa informasi jumlah (voucher acuan)")
        r += 1
        r = _write_table(
            ws6, r, ["Voucher No.", "Tanggal", "Deskripsi", "Nominal (Rp)"],
            teh_no_qty_df[["voucher", "tanggal_str", "deskripsi", "debit"]].values.tolist(),
            currency_cols=[4],
        )

    # --- Sheet 7: Mini Training & Jumsih ---
    ws7 = wb.create_sheet("Mini Training & Jumsih")
    mt_jum = alloc.groupby("bulan_str").agg(
        Mini_Training=("biaya_mini_training", "sum"),
        Jumsih=("biaya_jumsih", "sum"),
    ).reset_index()
    mt_jum["Bulan - Tahun"] = mt_jum["bulan_str"].apply(fmt_bulan)
    r = _write_title(ws7, 1, "Rincian biaya Mini Training dan Jumsih per bulan")
    r += 1
    r = _write_table(
        ws7, r,
        ["Bulan - Tahun", "Mini Training (Rp)", "Jumsih (Rp)"],
        mt_jum[["Bulan - Tahun", "Mini_Training", "Jumsih"]].round(0).values.tolist(),
        currency_cols=[2, 3],
    )
    r += 1
    ws7.cell(row=r, column=1,
             value="Catatan: Idealnya setiap bulan hanya ada salah satu (Mini Training atau Jumsih), tidak keduanya."
             ).font = BODY_FONT

    # --- Sheet 8: Lainnya ---
    ws8 = wb.create_sheet("Lainnya")
    # Transaksi yang tidak masuk kategori manapun (bukan galon, beras, gula/ktg, jumsih, mini training, bukber)
    mask_lainnya = (
        (alloc["galon_baru"] == 0) &
        (alloc["isi_ulang"] == 0) &
        (alloc["beras_kg"] == 0) &
        (~alloc["beras_no_qty"]) &
        (~alloc["ktg"]) &
        (~alloc["jumsih"]) &
        (~alloc["mini_training"]) &
        (~alloc["bukber"])
    )
    lainnya_df = alloc[mask_lainnya][["voucher", "tanggal", "deskripsi", "debit"]].copy()
    lainnya_df["tanggal"] = lainnya_df["tanggal"].dt.strftime("%d/%m/%Y")
    r = _write_title(ws8, 1, "Transaksi 'Lainnya' (tidak masuk kategori manapun)")
    r += 1
    ws8.cell(row=r, column=1,
             value="Daftar transaksi yang tidak terdeteksi sebagai galon, beras, gula, kopi, teh, jumsih, mini training, atau bukber."
             ).font = BODY_FONT
    r += 2
    if lainnya_df.empty:
        ws8.cell(row=r, column=1, value="Semua transaksi sudah masuk salah satu kategori.").font = BODY_FONT
    else:
        r = _write_table(
            ws8, r, ["Voucher No.", "Tanggal", "Deskripsi", "Nominal (Rp)"],
            lainnya_df[["voucher", "tanggal", "deskripsi", "debit"]].values.tolist(),
            currency_cols=[4],
        )

    # --- Sheet 9: All Transactions ---
    ws9 = wb.create_sheet("All Transactions")

    BULAN_ID_AT = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember",
    }

    at = alloc.copy()
    # Bulan - Tahun label: "Januari 2026"
    at["bulan_label"] = at["tanggal"].apply(
        lambda d: f"{BULAN_ID_AT[d.month]} {d.year}"
    )
    at["trans_date_str"] = at["tanggal"].dt.strftime("%d/%m/%Y")
    at["entry_date_str"] = at["entry_date"].apply(
        lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) else ""
    ) if "entry_date" in at.columns else ""
    # DIFF = selisih hari antara Entry Date dan Trans Date
    if "entry_date" in at.columns:
        at["diff_days"] = (at["entry_date"] - at["tanggal"]).dt.days.fillna("").astype(object)
        at.loc[at["entry_date"].isna(), "diff_days"] = ""
    else:
        at["diff_days"] = ""

    credit_col = at["credit"] if "credit" in at.columns else pd.Series(0.0, index=at.index)

    at_rows = []
    for _, row in at.iterrows():
        at_rows.append([
            row["bulan_label"],
            row["voucher"],
            row["trans_date_str"],
            row["entry_date_str"],
            row["deskripsi"],
            row["debit"],
            float(credit_col[row.name]) if "credit" in at.columns else 0.0,
            row["diff_days"],
        ])

    r = _write_title(ws9, 1, "Semua Transaksi (Amount in Base CCY)")
    r += 1
    r = _write_table(
        ws9, r,
        ["Bulan - Tahun", "Voucher No.", "Trans. Date", "Entry Date",
         "Description", "Debit", "Credit", "Diff (Hari)"],
        at_rows,
        currency_cols=[6, 7],
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =========================================================
# 4. STREAMLIT UI
# =========================================================

st.title("Breakdown Biaya Konsumsi Bulanan")
st.write(
    "Upload laporan transaksi akun Biaya Konsumsi (format tabel VOUCHER NO. / "
    "TRANS. DATE / ENTRY DATE / DESCRIPTION / DEBIT / CREDIT, atau format lama "
    "Transaction Listing By Accounts). Aplikasi ini memecah transaksi per bulan "
    "menjadi: kopi/gula/teh, beras, galon aqua asli (beli baru), isi ulang galon, "
    "serta jumsih dan mini training."
)

uploaded = st.file_uploader("Upload file laporan (.xls, .xlsx, atau .csv)", type=["xls", "xlsx", "csv"])

if uploaded is None:
    st.info("Silakan upload file laporan konsumsi untuk mulai.")
    st.stop()

raw = load_raw_table(uploaded)
if raw is None or raw.empty:
    st.error("File tidak bisa dibaca. Pastikan formatnya berupa tabel dengan kolom VOUCHER NO., TRANS. DATE, ENTRY DATE, DESCRIPTION, DEBIT, CREDIT.")
    st.stop()

trans = parse_report(raw)
if trans.empty:
    st.error("Tidak ada baris transaksi yang terdeteksi. Cek kembali format file (pastikan ada header VOUCHER NO. dan DEBIT).")
    st.stop()

st.write(
    f"Terbaca {len(trans)} transaksi, periode "
    f"{trans.tanggal.min().strftime('%d/%m/%Y')} sampai {trans.tanggal.max().strftime('%d/%m/%Y')}."
)

cat = categorize(trans)

st.sidebar.header("Pengaturan")

min_d, max_d = cat.tanggal.min().date(), cat.tanggal.max().date()
date_range = st.sidebar.date_input("Rentang tanggal", value=(min_d, max_d), min_value=min_d, max_value=max_d)
if isinstance(date_range, tuple) and len(date_range) == 2:
    start_d, end_d = date_range
    cat = cat[(cat.tanggal.dt.date >= start_d) & (cat.tanggal.dt.date <= end_d)]

auto_beras, auto_galon, auto_isi_ulang = estimate_unit_prices(cat)

st.sidebar.subheader("Estimasi harga satuan")
st.sidebar.write("Dihitung otomatis dari transaksi yang isinya cuma satu jenis item. Bisa diubah manual di bawah.")
harga_beras = st.sidebar.number_input("Harga per kg beras (Rp)", value=float(round(auto_beras)), step=100.0)
harga_galon = st.sidebar.number_input("Harga per galon aqua asli baru (Rp)", value=float(round(auto_galon)), step=500.0)
harga_isi_ulang = st.sidebar.number_input("Harga per isi ulang galon (Rp)", value=float(round(auto_isi_ulang)), step=500.0)

alloc = allocate_costs(cat, harga_beras, harga_galon, harga_isi_ulang)

st.divider()


def fmt_rp(x):
    return f"Rp {x:,.0f}".replace(",", ".")


alloc = alloc.assign(bulan_str=alloc["bulan"].astype(str))

# Hitung mask lainnya sekali, pakai di tab dan Excel
mask_lainnya_ui = (
    (alloc["galon_baru"] == 0) & (alloc["isi_ulang"] == 0) &
    (alloc["beras_kg"] == 0) & (~alloc["beras_no_qty"]) &
    (~alloc["ktg"]) & (~alloc["jumsih"]) &
    (~alloc["mini_training"]) & (~alloc["bukber"])
)
alloc["debit_lainnya"] = alloc["debit"].where(mask_lainnya_ui, 0)

# Tombol download di atas tab
excel_bytes = build_excel(alloc, harga_beras, harga_galon, harga_isi_ulang)
st.download_button(
    "⬇️ Download Rincian Excel",
    excel_bytes,
    "breakdown_konsumsi_bulanan.xlsx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.divider()

tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "Kopi/Gula/Teh", "Beras", "Galon Baru", "Isi Ulang Galon",
    "Jumsih & Mini Training", "Lainnya", "Ringkasan",
])

with tab1:
    st.write("Estimasi total belanja kopi, gula, dan teh per bulan.")
    g = alloc.groupby("bulan_str").agg(
        Total_Rp=("biaya_ktg", "sum"),
        Qty_Kopi=("kopi_qty", "sum"),
        Qty_Teh=("teh_qty", "sum"),
    ).reset_index()
    g.columns = ["Bulan", "Total (Rp)", "Qty Kopi (bungkus)", "Qty Teh (kantung/pcs)"]
    st.bar_chart(g.set_index("Bulan")["Total (Rp)"])
    disp1 = g.copy()
    disp1["Total (Rp)"] = disp1["Total (Rp)"].map(fmt_rp)
    st.dataframe(disp1, use_container_width=True, hide_index=True)
    st.caption(
        "Kolom 'Total (Rp)' adalah estimasi (nominal nota dikurangi estimasi beras/galon). "
        "Qty kopi/teh hanya terhitung jika deskripsi mencantumkan jumlah secara eksplisit."
    )

with tab2:
    st.write("Total kg beras dan estimasi biaya per bulan.")
    g = alloc.groupby("bulan_str").agg(Total_Kg=("beras_kg", "sum"), Total_Biaya=("biaya_beras", "sum")).reset_index()
    g.columns = ["Bulan", "Total Kg", "Estimasi Biaya (Rp)"]
    c1, c2 = st.columns(2)
    c1.bar_chart(g.set_index("Bulan")["Total Kg"])
    c2.bar_chart(g.set_index("Bulan")["Estimasi Biaya (Rp)"])
    st.dataframe(g.assign(**{"Estimasi Biaya (Rp)": g["Estimasi Biaya (Rp)"].map(fmt_rp)}),
                 use_container_width=True, hide_index=True)
    st.write(f"Estimasi harga per kg yang dipakai: {fmt_rp(harga_beras)}")

    missing = alloc[alloc.beras_no_qty][["voucher", "tanggal", "deskripsi", "debit"]]
    if len(missing):
        st.write("Transaksi beras yang tidak tercatat jumlah kg-nya:")
        m = missing.copy()
        m["tanggal"] = m["tanggal"].dt.strftime("%d/%m/%Y")
        m["debit"] = m["debit"].map(fmt_rp)
        m.columns = ["Voucher No.", "Tanggal", "Deskripsi", "Nominal"]
        st.dataframe(m, use_container_width=True, hide_index=True)

with tab3:
    st.write("Total galon aqua asli (beli baru) dan estimasi biaya per bulan.")
    g = alloc.groupby("bulan_str").agg(Total_Galon=("galon_baru", "sum"), Total_Biaya=("biaya_galon_baru", "sum")).reset_index()
    g.columns = ["Bulan", "Total Galon", "Estimasi Biaya (Rp)"]
    c1, c2 = st.columns(2)
    c1.bar_chart(g.set_index("Bulan")["Total Galon"])
    c2.bar_chart(g.set_index("Bulan")["Estimasi Biaya (Rp)"])
    st.dataframe(g.assign(**{"Estimasi Biaya (Rp)": g["Estimasi Biaya (Rp)"].map(fmt_rp)}),
                 use_container_width=True, hide_index=True)
    st.write(f"Estimasi harga per galon yang dipakai: {fmt_rp(harga_galon)}")

with tab4:
    st.write("Total isi ulang galon dan estimasi biaya per bulan.")
    g = alloc.groupby("bulan_str").agg(Total_Kali=("isi_ulang", "sum"), Total_Biaya=("biaya_isi_ulang", "sum")).reset_index()
    g.columns = ["Bulan", "Total Kali Isi Ulang", "Estimasi Biaya (Rp)"]
    c1, c2 = st.columns(2)
    c1.bar_chart(g.set_index("Bulan")["Total Kali Isi Ulang"])
    c2.bar_chart(g.set_index("Bulan")["Estimasi Biaya (Rp)"])
    st.dataframe(g.assign(**{"Estimasi Biaya (Rp)": g["Estimasi Biaya (Rp)"].map(fmt_rp)}),
                 use_container_width=True, hide_index=True)
    st.write(f"Estimasi harga per isi ulang yang dipakai: {fmt_rp(harga_isi_ulang)}")

with tab5:
    st.write("Biaya jumsih vs mini training per bulan. Idealnya hanya salah satu yang ada tiap bulan.")
    g = alloc.groupby("bulan_str").agg(
        Jumsih=("biaya_jumsih", "sum"),
        Mini_Training=("biaya_mini_training", "sum"),
    ).reset_index()
    g.columns = ["Bulan", "Jumsih (Rp)", "Mini Training (Rp)"]
    st.bar_chart(g.set_index("Bulan")[["Jumsih (Rp)", "Mini Training (Rp)"]])
    disp = g.copy()
    for col in ["Jumsih (Rp)", "Mini Training (Rp)"]:
        disp[col] = disp[col].map(fmt_rp)
    st.dataframe(disp, use_container_width=True, hide_index=True)

with tab6:
    st.write("Transaksi yang tidak terdeteksi masuk kategori manapun (bukan galon, beras, gula, kopi, teh, jumsih, mini training, atau bukber).")
    lainnya_ui = alloc[mask_lainnya_ui].copy()
    if lainnya_ui.empty:
        st.info("Semua transaksi sudah teridentifikasi kategorinya.")
    else:
        g_l = lainnya_ui.groupby("bulan_str").agg(
            Jumlah_Transaksi=("debit", "count"),
            Total_Nominal=("debit", "sum"),
        ).reset_index()
        g_l.columns = ["Bulan", "Jumlah Transaksi", "Total Nominal (Rp)"]
        st.bar_chart(g_l.set_index("Bulan")["Total Nominal (Rp)"])
        g_l["Total Nominal (Rp)"] = g_l["Total Nominal (Rp)"].map(fmt_rp)
        st.dataframe(g_l, use_container_width=True, hide_index=True)
        st.write("Detail transaksi:")
        det = lainnya_ui[["voucher", "tanggal", "deskripsi", "debit"]].copy()
        det["tanggal"] = det["tanggal"].dt.strftime("%d/%m/%Y")
        det["debit"] = det["debit"].map(fmt_rp)
        det.columns = ["Voucher No.", "Tanggal", "Deskripsi", "Nominal"]
        st.dataframe(det, use_container_width=True, hide_index=True)

with tab7:
    st.write("Ringkasan seluruh kategori per bulan.")
    ring = alloc.groupby("bulan_str").agg(
        Kopi_Gula_Teh=("biaya_ktg", "sum"),
        Qty_Kopi=("kopi_qty", "sum"),
        Qty_Teh=("teh_qty", "sum"),
        Beras_Rp=("biaya_beras", "sum"),
        Beras_Kg=("beras_kg", "sum"),
        Galon_Rp=("biaya_galon_baru", "sum"),
        Galon_Qty=("galon_baru", "sum"),
        IsiUlang_Rp=("biaya_isi_ulang", "sum"),
        IsiUlang_Qty=("isi_ulang", "sum"),
        Jumsih=("biaya_jumsih", "sum"),
        Mini_Training=("biaya_mini_training", "sum"),
        Lainnya=("debit_lainnya", "sum"),
        Total_Debit_Asli=("debit", "sum"),
    ).reset_index()

    disp = ring[["bulan_str",
                 "Kopi_Gula_Teh", "Qty_Kopi", "Qty_Teh",
                 "Beras_Rp", "Beras_Kg",
                 "Galon_Rp", "Galon_Qty",
                 "IsiUlang_Rp", "IsiUlang_Qty",
                 "Jumsih", "Mini_Training",
                 "Lainnya", "Total_Debit_Asli"]].copy()
    disp.columns = [
        "Bulan",
        "KGT (Rp)", "Kopi (bungkus)", "Teh (kantung/pcs)",
        "Beras (Rp)", "Beras (Kg)",
        "Galon Baru (Rp)", "Galon Baru (Qty)",
        "Isi Ulang (Rp)", "Isi Ulang (Qty)",
        "Jumsih (Rp)", "Mini Training (Rp)",
        "Lainnya (Rp)", "Total Debit Asli (Rp)",
    ]
    for col in ["KGT (Rp)", "Beras (Rp)", "Galon Baru (Rp)", "Isi Ulang (Rp)",
                "Jumsih (Rp)", "Mini Training (Rp)", "Lainnya (Rp)", "Total Debit Asli (Rp)"]:
        disp[col] = disp[col].map(fmt_rp)
    st.dataframe(disp, use_container_width=True, hide_index=True)

    st.caption(
        "'Lainnya' = total nominal debit transaksi yang tidak masuk kategori manapun. "
        "Qty kopi/teh hanya terhitung jika deskripsi mencantumkan jumlah eksplisit."
    )

st.divider()
with st.expander("Lihat data transaksi mentah yang sudah dikategorikan"):
    show_cols = ["voucher", "tanggal", "deskripsi", "debit", "beras_kg", "gula_kg",
                 "galon_baru", "isi_ulang", "kopi", "kopi_qty", "teh", "teh_qty",
                 "ktg", "jumsih", "mini_training", "bukber"]
    st.dataframe(alloc[show_cols], use_container_width=True, hide_index=True)
